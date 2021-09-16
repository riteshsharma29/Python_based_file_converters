import pandas as pd
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import load_workbook
import os
import sys

filename = sys.argv[1]

def csv_2_excel(filename):

    try:
        # Create Excel Workbook
        DATA_File = os.path.join(os.getcwd(), str(filename) + '.xlsx')
        workbook = xlsxwriter.Workbook(DATA_File)
        worksheet = workbook.add_worksheet()
        workbook.close()

        #  Load excel Workbook using openpyxl
        book = load_workbook(DATA_File)
        writer = ExcelWriter(DATA_File, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # convert csv into a dataframe
        df = pd.read_csv(filename)
        df.to_excel(writer, sheet_name="data", index=False)

        # remove default empty first sheet
        if len(book.sheetnames) > 1:
            first_sheet = book['Sheet1']
            book.remove(first_sheet)
        writer.save()

    except Exception as e:
        print("")
        print("")
        print("csv to xlsx conversions errors are :")
        print("")
        print("")
        print(e)
        print("")
        print("")

csv_2_excel(filename)